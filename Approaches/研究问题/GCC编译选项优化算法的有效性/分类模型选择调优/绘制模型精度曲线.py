from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import random
import numpy as np
from openpyxl import load_workbook  # 导入openpyxl

import matplotlib.pyplot as plt
import pandas as pd
import random

# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data


class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

#返回均值列表和标准差列表
def get_meanList_and_stdList(case_list):
    mean_list=[]
    std_list=[]
    for i in range(len(case_list[0])):
        temp=[]
        for j in range(len(case_list)):
            temp.append(case_list[j][i])
        mean_list.append(np.mean(temp))
        std_list.append(np.std(temp))
    return  mean_list,std_list

def list_trun(case_list):

    ret_list=[]
    for case_one_list in case_list:
        temp = []
        for j in case_one_list:
            temp.append(round((j*2)/100.0,3))
        ret_list.append(temp)
    return  ret_list


if __name__ == '__main__':
    casename="automotive_susan_s"
    data=doExcel("各模型精度.xlsx", casename).get_data()
    Logic=[]
    RF=[]
    Tree = []
    SVM=[]
    for i in range(4):
        Logic.append(list(eval(data[i]['逻辑回归'])))
    for i in range(4):
        RF.append(list(eval(data[i]['随机森林'])))
    for i in range(4):
        Tree.append(list(eval(data[i]['决策树'])))
    for i in range(4):
        SVM.append(list(eval(data[i]['支持向量机'])))

    trun_Logic = list_trun(Logic)
    trun_RF = list_trun(RF)
    trun_Tree = list_trun(Tree)
    trun_SVM = list_trun(SVM)


    mean_Logic,std_Logic = get_meanList_and_stdList(trun_Logic)
    mean_Logic,std_Logic = np.array(mean_Logic),np.array(std_Logic)

    mean_RF,std_RF = get_meanList_and_stdList(trun_RF)
    mean_RF, std_RF = np.array(mean_RF), np.array(std_RF)

    mean_Tree,std_Tree = get_meanList_and_stdList(trun_Tree)
    mean_Tree, std_Tree = np.array(mean_Tree), np.array(std_Tree)

    mean_SVM,std_SVM = get_meanList_and_stdList(trun_SVM)
    mean_SVM, std_SVM = np.array(mean_SVM), np.array(std_SVM)

    x = []
    for i in range(10):
        x.append(i+1)
    x=np.array(x)

    plt.plot(x, mean_Logic , 'r-', label='逻辑回归')
    plt.fill_between(x, mean_Logic - std_Logic, mean_Logic + std_Logic, color='b', alpha=0.2)

    plt.plot(x, mean_RF, 'b-', label='随机森林')
    plt.fill_between(x, mean_RF - std_RF, mean_RF + std_RF, color='r', alpha=0.2)

    plt.plot(x, mean_Tree, 'y-', label='决策树')
    plt.fill_between(x, mean_Tree - std_Tree, mean_Tree + std_Tree, color='r', alpha=0.2)

    plt.plot(x,mean_SVM, 'g-', label='支持向量机')
    plt.fill_between(x, mean_SVM - std_SVM, mean_SVM + std_SVM, color='r', alpha=0.2)
    plt.ylim(0.6,1)
    plt.xlim(1, 10)
    plt.xticks([1, 2, 3, 4, 5, 6, 7, 8,9,10])
    plt.xlabel("迭代次数" ,fontsize=12)
    plt.ylabel("预测准确率" ,fontsize=12)

    plt.legend()
    plt.show()





