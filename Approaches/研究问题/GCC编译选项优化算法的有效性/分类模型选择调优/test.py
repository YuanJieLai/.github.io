from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import random
import numpy as np
from openpyxl import load_workbook  # 导入openpyxl

import matplotlib.pyplot as plt
import pandas as pd
import random

# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data


class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

#返回均值列表和标准差列表
def get_meanList_and_stdList(case_list):
    mean_list=[]
    std_list=[]
    for i in range(len(case_list[0])):
        temp=[]
        for j in range(len(case_list)):
            temp.append(case_list[j][i])
        mean_list.append(np.mean(temp))
        std_list.append(np.std(temp))
    return  mean_list,std_list

def list_trun(case_list):

    ret_list=[]
    for case_one_list in case_list:
        temp = []
        for j in case_one_list:
            temp.append(round((j*2)/100.0,3))
        ret_list.append(temp)
    return  ret_list


if __name__ == '__main__':
    casename="consumer_jpeg_c"
    data=doExcel("各模型精度.xlsx", casename).get_data()
    Logic=[]
    RF=[]
    Tree = []
    SVM=[]
    for i in range(4):
        Logic.append(list(eval(data[i]['逻辑回归'])))
    for i in range(4):
        RF.append(list(eval(data[i]['随机森林'])))
    for i in range(4):
        Tree.append(list(eval(data[i]['决策树'])))
    for i in range(4):
        SVM.append(list(eval(data[i]['支持向量机'])))



    print(Logic)

    avg_Logic=[]
    for i in range(len(Logic[0])):
        data_sum=0.0
        for j in range(len(Logic)):
            data_sum = data_sum + Logic[j][i]
        avg_Logic.append(round(data_sum/4.0,3))

    avg_RF = []
    for i in range(len(RF[0])):
        data_sum = 0.0
        for j in range(len(RF)):
            data_sum = data_sum + RF[j][i]
        avg_RF.append(round(data_sum / 4.0, 3))

    avg_Tree = []
    for i in range(len(Tree[0])):
        data_sum = 0.0
        for j in range(len(Tree)):
            data_sum = data_sum + Tree[j][i]
        avg_Tree.append(round(data_sum / 4.0, 3))

    avg_SVM = []
    for i in range(len(SVM[0])):
        data_sum = 0.0
        for j in range(len(SVM)):
            data_sum = data_sum + SVM[j][i]
        avg_SVM.append(round(data_sum / 4.0, 3))

    print("************",casename,"************")

    print(avg_Logic)
    print(avg_RF)
    print(avg_Tree)
    print(avg_SVM)

#计算总的平均值
    total_avg=[]

    total_sum=0.0
    for i in range(len(avg_Logic)):
        total_sum=total_sum+avg_Logic[i]
    total_avg.append(total_sum/10.0)

    total_sum=0.0
    for i in range(len(avg_RF)):
        total_sum=total_sum+avg_RF[i]
    total_avg.append(total_sum/10.0)

    total_sum=0.0
    for i in range(len(avg_Tree)):
        total_sum=total_sum+avg_Tree[i]
    total_avg.append(total_sum/10.0)

    total_sum=0.0
    for i in range(len(avg_SVM)):
        total_sum=total_sum+avg_SVM[i]
    total_avg.append(total_sum/10.0)

    print("平均准确率",total_avg)











