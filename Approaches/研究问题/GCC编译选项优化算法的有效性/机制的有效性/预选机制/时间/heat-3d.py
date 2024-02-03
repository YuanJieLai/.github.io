from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

if __name__ == '__main__':
    data=doExcel("有无机制最优解表.xlsx", "Sheet2").get_data()


    print(data)

    Our_2 =list(eval(data[0]['automotive_susan_s']))
    Our_3 =list(eval(data[0]['consumer_jpeg_c']))
    Our_1 = list(eval(data[0]['3mm']))
    Our_4 = list(eval(data[0]['heat-3d']))
    Our_5 =list(eval(data[0]['telecom_adpcm_c']))

    s_2 =list(eval(data[1]['automotive_susan_s']))
    s_3 =list(eval(data[1]['consumer_jpeg_c']))
    s_1 = list(eval(data[1]['3mm']))
    s_4 = list(eval(data[1]['heat-3d']))
    s_5 =list(eval(data[1]['telecom_adpcm_c']))


    df11 = {'C1(SRA-BM)': Our_4,
             }
    df21 = {'C1(CBOS)': s_4,
             }
    df1 = pd.DataFrame(df11)
    df2 = pd.DataFrame(df21)

    # df1.plot.box(title="", showfliers=None)
    plt.grid(linestyle="--", alpha=0.3)
    # df2.plot.box(title="", showfliers=None)
    plt.tick_params(labelsize=6)
    # 25
    plt.xticks(rotation=0)
    # plt.savefig('./results_imgs.png', bbox_inches='tight')
    font2 = {'family': 'Times New Roman',
             'weight': 'normal',
             'size': 20,
             }
    plt.boxplot(x=df1,

                #     patch_artist=True,  # 要求用自定义颜色填充盒形图，默认白色填充

                showmeans=True,  # 以点的形式显示均值
                positions=[1],
                #  boxprops={'color': 'black', 'facecolor': '#9999ff'},  # 设置箱体属性，填充色和边框色

                flierprops={'marker': 'o', 'markerfacecolor': 'red', 'color': 'black'},  # 设置异常值属性，点的形状、填充色和边框色

                meanprops={'marker': 'D', 'markerfacecolor': 'indianred'},  # 设置均值点的属性，点的形状、填充色

                medianprops={'linestyle': '--', 'color': 'red'})  # 设置中位数线的属性，线的类型和颜色

    plt.boxplot(x=df2,

                patch_artist=True,  # 要求用自定义颜色填充盒形图，默认白色填充

                showmeans=True,  # 以点的形式显示均值
                positions=[2],
                boxprops={'color': 'black', 'facecolor': '#9999ff'},  # 设置箱体属性，填充色和边框色

                flierprops={'marker': 'o', 'markerfacecolor': 'blue', 'color': 'black'},  # 设置异常值属性，点的形状、填充色和边框色

                meanprops={'marker': 'd', 'markerfacecolor': 'indianred'},  # 设置均值点的属性，点的形状、填充色

                medianprops={'linestyle': '--', 'color': 'red'})  # 设置中位数线的属性，线的类型和颜色

    plt.xticks([1, 2],
               ['OSOA(heat-3d)', 'OSOA-s(heat-3d)'],fontsize=13)
    plt.yticks( fontsize=8)
    plt.ylabel("$Run time$", fontsize=13)
    # plt.savefig('./recallb1.svg', dpi=1200, format='svg')
    plt.show()

    # plt.savefig('./results_imgs.png', bbox_inches='tight')






