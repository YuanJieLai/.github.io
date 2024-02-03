from scipy.stats import wilcoxon
from numpy import median
from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
from math import sqrt
import matplotlib.pyplot as plt
import pandas as pd

# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data
# 两个样本的数据
# sample1 = [1, 2, 3, 4, 5]
# sample2 = [7, 8, 9, 11,12]

data = doExcel("各案例指标值.xlsx", "Sheet1").get_data()

Our_C1 = list(eval(data[0]['automotive_bitcount']))
Our_C2 = list(eval(data[0]['automotive_susan_c']))
Our_C3 = list(eval(data[0]['automotive_susan_e']))
Our_C4 = list(eval(data[0]['automotive_susan_s']))
Our_C5 = list(eval(data[0]['bzip2e']))
Our_C6 = list(eval(data[0]['consumer_jpeg_c']))
Our_C7 = list(eval(data[0]['consumer_tiff2rgba']))
Our_C8 = list(eval(data[0]['office_rsynth']))
Our_C9 = list(eval(data[0]['security_sha']))
Our_C10 = list(eval(data[0]['telecom_adpcm_c']))

CB_C1 = list(eval(data[3]['automotive_bitcount']))
CB_C2 = list(eval(data[3]['automotive_susan_c']))
CB_C3 = list(eval(data[3]['automotive_susan_e']))
CB_C4 = list(eval(data[3]['automotive_susan_s']))
CB_C5 = list(eval(data[3]['bzip2e']))
CB_C6 = list(eval(data[3]['consumer_jpeg_c']))
CB_C7 = list(eval(data[3]['consumer_tiff2rgba']))
CB_C8 = list(eval(data[3]['office_rsynth']))
CB_C9 = list(eval(data[3]['security_sha']))
CB_C10 = list(eval(data[3]['telecom_adpcm_c']))

Our_P1 = list(eval(data[0]['correlation']))
Our_P2 = list(eval(data[0]['covariance']))
Our_P3 = list(eval(data[0]['symm']))
Our_P4 = list(eval(data[0]['2mm']))
Our_P5 = list(eval(data[0]['3mm']))
Our_P6 = list(eval(data[0]['lu']))
Our_P7 = list(eval(data[0]['cholesky']))
Our_P8 = list(eval(data[0]['nussinov']))
Our_P9 = list(eval(data[0]['heat-3d']))
Our_P10 = list(eval(data[0]['jacobi-2d']))

CB_P1 = list(eval(data[3]['correlation']))
CB_P2 = list(eval(data[3]['covariance']))
CB_P3 = list(eval(data[3]['symm']))
CB_P4 = list(eval(data[3]['2mm']))
CB_P5 = list(eval(data[3]['3mm']))
CB_P6 = list(eval(data[3]['lu']))
CB_P7 = list(eval(data[3]['cholesky']))
CB_P8 = list(eval(data[3]['nussinov']))
CB_P9 = list(eval(data[3]['heat-3d']))
CB_P10 = list(eval(data[3]['jacobi-2d']))





sample1 =Our_P10
sample2 =CB_P10

print(sample1)
print(sample2)


wilcoxon_result = wilcoxon(sample1, sample2)
p_value = wilcoxon_result.pvalue
effect_size = wilcoxon_result.statistic / (len(sample1) * len(sample2))

# 输出结果
print("p-value:", p_value)
# 计算样本差异的平均值
mean_diff = sum(np.array(sample1) - np.array(sample2)) / len(sample1)
# 计算样本差异的标准差
var_diff = sum((np.array(sample1) - np.array(sample2) - mean_diff) ** 2) / (len(sample1) - 1)
std_diff = sqrt(var_diff)

# 计算Cohen's d效应值
cohens_d = mean_diff / std_diff
print("Cohen's d效应值为：", cohens_d)





