from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

if __name__ == '__main__':
    data=doExcel("各案例指标值.xlsx", "Sheet1").get_data()


    Our_C1 =list(eval(data[0]['automotive_bitcount']))
    Our_C2 =list(eval(data[0]['automotive_susan_c']))
    Our_C3 =list(eval(data[0]['automotive_susan_e']))
    Our_C4 =list(eval(data[0]['automotive_susan_s']))
    Our_C5 =list(eval(data[0]['bzip2e']))
    Our_C6 =list(eval(data[0]['consumer_jpeg_c']))
    Our_C7 =list(eval(data[0]['consumer_tiff2rgba']))
    Our_C8 =list(eval(data[0]['office_rsynth']))
    Our_C9 =list(eval(data[0]['security_sha']))
    Our_C10 =list(eval(data[0]['telecom_adpcm_c']))

    CB_C1 =list(eval(data[3]['automotive_bitcount']))
    CB_C2 =list(eval(data[3]['automotive_susan_c']))
    CB_C3 =list(eval(data[3]['automotive_susan_e']))
    CB_C4 =list(eval(data[3]['automotive_susan_s']))
    CB_C5 =list(eval(data[3]['bzip2e']))
    CB_C6 =list(eval(data[3]['consumer_jpeg_c']))
    CB_C7 =list(eval(data[3]['consumer_tiff2rgba']))
    CB_C8 =list(eval(data[3]['office_rsynth']))
    CB_C9 =list(eval(data[3]['security_sha']))
    CB_C10 =list(eval(data[3]['telecom_adpcm_c']))


    Our_P1 =list(eval(data[0]['correlation']))
    Our_P2 =list(eval(data[0]['covariance']))
    Our_P3 =list(eval(data[0]['symm']))
    Our_P4 =list(eval(data[0]['2mm']))
    Our_P5 =list(eval(data[0]['3mm']))
    Our_P6 =list(eval(data[0]['lu']))
    Our_P7 =list(eval(data[0]['cholesky']))
    Our_P8 =list(eval(data[0]['nussinov']))
    Our_P9 =list(eval(data[0]['heat-3d']))
    Our_P10 =list(eval(data[0]['jacobi-2d']))


    CB_P1 =list(eval(data[3]['correlation']))
    CB_P2 =list(eval(data[3]['covariance']))
    CB_P3 =list(eval(data[3]['symm']))
    CB_P4 =list(eval(data[3]['2mm']))
    CB_P5 =list(eval(data[3]['3mm']))
    CB_P6 =list(eval(data[3]['lu']))
    CB_P7 =list(eval(data[3]['cholesky']))
    CB_P8 =list(eval(data[3]['nussinov']))
    CB_P9 =list(eval(data[3]['heat-3d']))
    CB_P10 =list(eval(data[3]['jacobi-2d']))





    dfOSA = {'C1(SRA-BM)': Our_C1,
             'C2(SRA-BM)': Our_C2,
             'C3(SRA-BM)': Our_C3,
             'C4(SRA-BM)': Our_C4,
             'C5(SRA-BM)': Our_C5,
             'C6(SRA-BM)': Our_C6,
             'C7(SRA-BM)': Our_C7,
             'C8(SRA-BM)': Our_C8,
             'C9(SRA-BM)': Our_C9,
             'C10(SRA-BM)': Our_C10,
             }
    dfC4O = {'C1(CBOS)': CB_C1,
             'C2(CBOS)': CB_C2,
             'C3(CBOS)': CB_C3,
             'C4(CBOS)': CB_C4,
             'C5(CBOS)': CB_C5,
             'C6(CBOS)': CB_C6,
             'C7(CBOS)': CB_C7,
             'C8(CBOS)': CB_C8,
             'C9(CBOS)': CB_C9,
             'C10(CBOS)': CB_C10,
             }
    df1 = pd.DataFrame(dfOSA)
    df2 = pd.DataFrame(dfC4O)
    # df1.plot.box(title="", showfliers=None)
    plt.grid(linestyle="--", alpha=0.3)
    # df2.plot.box(title="", showfliers=None)
    plt.tick_params(labelsize=6)
    # 25
    plt.xticks(rotation=20)
    # plt.savefig('./results_imgs.png', bbox_inches='tight')
    font2 = {'family': 'Times New Roman',
             'weight': 'normal',
             'size': 15,
             }
    plt.boxplot(x=df1,

                #     patch_artist=True,  # 要求用自定义颜色填充盒形图，默认白色填充

                showmeans=True,  # 以点的形式显示均值
                positions=[1, 3, 5, 7,9,11, 13,  15,  17, 19],
                #  boxprops={'color': 'black', 'facecolor': '#9999ff'},  # 设置箱体属性，填充色和边框色

                flierprops={'marker': 'o', 'markerfacecolor': 'red', 'color': 'black'},  # 设置异常值属性，点的形状、填充色和边框色

                meanprops={'marker': 'D', 'markerfacecolor': 'indianred'},  # 设置均值点的属性，点的形状、填充色

                medianprops={'linestyle': '--', 'color': 'red'})  # 设置中位数线的属性，线的类型和颜色

    plt.boxplot(x=df2,

                patch_artist=True,  # 要求用自定义颜色填充盒形图，默认白色填充

                showmeans=True,  # 以点的形式显示均值
                positions=[2, 4, 6, 8,10,12, 14, 16,  18,20],
                boxprops={'color': 'black', 'facecolor': '#9999ff'},  # 设置箱体属性，填充色和边框色

                flierprops={'marker': 'o', 'markerfacecolor': 'blue', 'color': 'black'},  # 设置异常值属性，点的形状、填充色和边框色

                meanprops={'marker': 'd', 'markerfacecolor': 'indianred'},  # 设置均值点的属性，点的形状、填充色

                medianprops={'linestyle': '--', 'color': 'red'})  # 设置中位数线的属性，线的类型和颜色

    plt.xticks([1, 2, 3, 4, 5, 6, 7, 8,9,10,11, 12, 13, 14, 15, 16, 17, 18,19,20],
               ['C1(OPA)', 'C1(CB)', 'C2(OPA)', 'C2(CB)', 'C3(OPA)',
                'C3(CB)', 'C4(OPA)', 'C4(CB)', 'C5(OPA)', 'C5(CB)',
                'C6(OPA)', 'C6(CB)', 'C7(OPA)', 'C7(CB)', 'C8(OPA)',
                'C8(CB)', 'C9(OPA)', 'C9(CB)', 'C10(OPA)', 'C10(CB)'])

    plt.ylabel("$Recall_{b}$", fontsize=20)
    #plt.savefig('./recallb1.svg', dpi=1200, format='svg')
    plt.show()

    #plt.savefig('./results_imgs.png', bbox_inches='tight')
