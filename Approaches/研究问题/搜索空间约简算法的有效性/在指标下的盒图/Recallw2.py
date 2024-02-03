from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd

# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

if __name__ == '__main__':
    data=doExcel("各案例指标值.xlsx", "Sheet1").get_data()

    Our_C1 =list(eval(data[1]['automotive_bitcount']))
    Our_C2 =list(eval(data[1]['automotive_susan_c']))
    Our_C3 =list(eval(data[1]['automotive_susan_e']))
    Our_C4 =list(eval(data[1]['automotive_susan_s']))
    Our_C5 =list(eval(data[1]['bzip2e']))
    Our_C6 =list(eval(data[1]['consumer_jpeg_c']))
    Our_C7 =list(eval(data[1]['consumer_tiff2rgba']))
    Our_C8 =list(eval(data[1]['office_rsynth']))
    Our_C9 =list(eval(data[1]['security_sha']))
    Our_C10 =list(eval(data[1]['telecom_adpcm_c']))

    CB_C1 =list(eval(data[4]['automotive_bitcount']))
    CB_C2 =list(eval(data[4]['automotive_susan_c']))
    CB_C3 =list(eval(data[4]['automotive_susan_e']))
    CB_C4 =list(eval(data[4]['automotive_susan_s']))
    CB_C5 =list(eval(data[4]['bzip2e']))
    CB_C6 =list(eval(data[4]['consumer_jpeg_c']))
    CB_C7 =list(eval(data[4]['consumer_tiff2rgba']))
    CB_C8 =list(eval(data[4]['office_rsynth']))
    CB_C9 =list(eval(data[4]['security_sha']))
    CB_C10 =list(eval(data[4]['telecom_adpcm_c']))


    Our_P1 =list(eval(data[1]['correlation']))
    Our_P2 =list(eval(data[1]['covariance']))
    Our_P3 =list(eval(data[1]['symm']))
    Our_P4 =list(eval(data[1]['2mm']))
    Our_P5 =list(eval(data[1]['3mm']))
    Our_P6 =list(eval(data[1]['lu']))
    Our_P7 =list(eval(data[1]['cholesky']))
    Our_P8 =list(eval(data[1]['nussinov']))
    Our_P9 =list(eval(data[1]['heat-3d']))
    Our_P10 =list(eval(data[1]['jacobi-2d']))


    CB_P1 =list(eval(data[4]['correlation']))
    CB_P2 =list(eval(data[4]['covariance']))
    CB_P3 =list(eval(data[4]['symm']))
    CB_P4 =list(eval(data[4]['2mm']))
    CB_P5 =list(eval(data[4]['3mm']))
    CB_P6 =list(eval(data[4]['lu']))
    CB_P7 =list(eval(data[4]['cholesky']))
    CB_P8 =list(eval(data[4]['nussinov']))
    CB_P9 =list(eval(data[4]['heat-3d']))
    CB_P10 =list(eval(data[4]['jacobi-2d']))





    dfOSA = {'C1(SRA-BM)': Our_P1,
             'C2(SRA-BM)': Our_P2,
             'C3(SRA-BM)': Our_P3,
             'C4(SRA-BM)': Our_P4,
             'C5(SRA-BM)': Our_P5,
             'C6(SRA-BM)': Our_P6,
             'C7(SRA-BM)': Our_P7,
             'C8(SRA-BM)': Our_P8,
             'C9(SRA-BM)': Our_P9,
             'C10(SRA-BM)': Our_P10,
             }
    dfC4O = {'C1(CBOS)': CB_P1,
             'C2(CBOS)': CB_P2,
             'C3(CBOS)': CB_P3,
             'C4(CBOS)': CB_P4,
             'C5(CBOS)': CB_P5,
             'C6(CBOS)': CB_P6,
             'C7(CBOS)': CB_P7,
             'C8(CBOS)': CB_P8,
             'C9(CBOS)': CB_P9,
             'C10(CBOS)': CB_P10,
             }
    df1 = pd.DataFrame(dfOSA)
    df2 = pd.DataFrame(dfC4O)
    # df1.plot.box(title="", showfliers=None)
    plt.grid(linestyle="--", alpha=0.3)
    # df2.plot.box(title="", showfliers=None)
    plt.tick_params(labelsize=6)
    # 25
    plt.xticks(rotation=20)
    # plt.savefig('./results_imgs.png', bbox_inches='tight')
    font2 = {'family': 'Times New Roman',
             'weight': 'normal',
             'size': 15,
             }
    plt.boxplot(x=df1,

                #     patch_artist=True,  # 要求用自定义颜色填充盒形图，默认白色填充

                showmeans=True,  # 以点的形式显示均值
                positions=[1, 3, 5, 7,9,11, 13,  15,  17, 19],
                #  boxprops={'color': 'black', 'facecolor': '#9999ff'},  # 设置箱体属性，填充色和边框色

                flierprops={'marker': 'o', 'markerfacecolor': 'red', 'color': 'black'},  # 设置异常值属性，点的形状、填充色和边框色

                meanprops={'marker': 'D', 'markerfacecolor': 'indianred'},  # 设置均值点的属性，点的形状、填充色

                medianprops={'linestyle': '--', 'color': 'red'})  # 设置中位数线的属性，线的类型和颜色

    plt.boxplot(x=df2,

                patch_artist=True,  # 要求用自定义颜色填充盒形图，默认白色填充

                showmeans=True,  # 以点的形式显示均值
                positions=[2, 4, 6, 8,10,12, 14, 16,  18,20],
                boxprops={'color': 'black', 'facecolor': '#9999ff'},  # 设置箱体属性，填充色和边框色

                flierprops={'marker': 'o', 'markerfacecolor': 'blue', 'color': 'black'},  # 设置异常值属性，点的形状、填充色和边框色

                meanprops={'marker': 'd', 'markerfacecolor': 'indianred'},  # 设置均值点的属性，点的形状、填充色

                medianprops={'linestyle': '--', 'color': 'red'})  # 设置中位数线的属性，线的类型和颜色

    plt.xticks([1, 2, 3, 4, 5, 6, 7, 8,9,10,11, 12, 13, 14, 15, 16, 17, 18,19,20],
               ['P1(OPA)', 'P1(CB)', 'P2(OPA)', 'P2(CB)', 'P3(OPA)',
                'P3(CB)', 'P4(OPA)', 'P4(CB)', 'P5(OPA)', 'P5(CB)',
                'P6(OPA)', 'P6(CB)', 'P7(OPA)', 'P7(CB)', 'P8(OPA)',
                'P8(CB)', 'P9(OPA)', 'P9(CB)', 'P10(OPA)', 'P10(CB)'])

    plt.ylabel("$Recall_{w}$", fontsize=20)
    #plt.savefig('./recallb1.svg', dpi=1200, format='svg')
    plt.show()

    #plt.savefig('./results_imgs.png', bbox_inches='tight')
