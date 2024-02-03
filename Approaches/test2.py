from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import random
# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
fuzzy=[71, 0.4, 0.3, 0.3, 0.3, 0.8, 0.3, 0.7, 0.3, 0.4, 0.3, 0.1, 0.2, 0.4, 0.3, 0.5, 0.2, 0.0, 0.4, 0.3, 0.2, 0.3, 0.5, 0.3, 0.3, 0.3, 0.3, 0.3, 0.0, 0.4, 0.3, 0.5, 0.4, 0.7, 0.3, 0.2, 0.4, 0.2, 0.4, 0.1, 0.8, 0.3, 0.3, 0.4, 0.2, 0.2, 0.4, 0.1, 0.2, 0.4, 0.4, 0.4, 0.2, 0.2, 0.6, 0.4, 0.4, 0.4, 0.2, 0.2, 0.2, 0.2, 0.4, 0.3, 0.3, 0.6, 0.3, 0.4, 0.6, 0.3, 0.3, 0.3]

class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

if __name__ == '__main__':
    print("**************start*******************")
    S = [71, 0.4, 0.3, 0.3, 0.3, 0.8, 0.3, 0.7, 0.3, 0.4, 0.3, 0.1, 0.2, 0.4, 0.3, 0.5, 0.2, 0.0, 0.4, 0.3, 0.2, 0.3,
         0.5, 0.3, 0.3, 0.3, 0.3, 0.3, 0.0, 0.4, 0.3, 0.5, 0.4, 0.7, 0.3, 0.2, 0.4, 0.2, 0.4, 0.1, 0.8, 0.3, 0.3, 0.4,
         0.2, 0.2, 0.4, 0.1, 0.2, 0.4, 0.4, 0.4, 0.2, 0.2, 0.6, 0.4, 0.4, 0.4, 0.2, 0.2, 0.2, 0.2, 0.4, 0.3, 0.3, 0.6,
         0.3, 0.4, 0.6, 0.3, 0.3, 0.3]
    NS = []
    NS.append(71)
    for i in range(71):
        hhh=1-S[i+1]
        NS.append(hhh)
    TP = 0.0
    FN = 0.0
    FP = 0.0
    TN = 0.0
    select_opt=[]
    notselect_opt = []
    select_opt.append(71)
    notselect_opt.append(71)
    for i in range(71):
        notselect_opt.append(0)
    for i in range(71):
        select_opt.append(1)
    for i in range(71):
        if (select_opt[i + 1] == 0):
            notselect_opt[i + 1] = 1

    for i in range(71):
        if (select_opt[i + 1] == 1):
            TP = TP + S[i + 1]
            FP = FP + NS[i + 1]
    for i in range(71):
        if (notselect_opt[i + 1] == 1):
            TN = TN + NS[i + 1]
            FN = FN + S[i + 1]

    CRecallb = TP / (TP + FN)
    CRecallw = TN / (TN + FP)
    CAccuracy = (TP + TN) / (TP + TN + FP + FN)

    print("TP=", TP, "FP=", FP, "TN=", TN, "FN=", FN)
    print("*************CRecallb=", CRecallb)
    print("*************CRecallw=", CRecallw)
    print("*************CAccuracy=", CAccuracy)











    print("**************end*******************")