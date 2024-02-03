from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import random
import xlwt

# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False


class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

#读取各个案例下的指标值
if __name__ == '__main__':
    print("**************start*******************")

    casename = "telecom_adpcm_c"

    recallb = []
    recallw = []
    accuracy = []
    for alg_temp_it in range(10):
        data = doExcel("原始序列结果总集.xlsx", casename+str(alg_temp_it+1)).get_data()

        recallb.append(round(float(data[22]['正向查全率']),3))
        recallw.append(round(float(data[22]['负向查全率']),3))
        accuracy.append(round(float(data[22]['精度']),3))
        print(data)

    print("casename=",casename)
    print("recallb=",recallb)
    print("recallw=",recallw)
    print("accuracy=",accuracy)