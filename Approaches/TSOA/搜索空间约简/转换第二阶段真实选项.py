from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import random
import xlwt

# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False


class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

#读取各个案例下的指标值
if __name__ == '__main__':
    print("**************start*******************")

    casename = "office_rsynth"


    for alg_temp_it in range(1):
        data = doExcel("原始序列结果总集.xlsx", casename+str(alg_temp_it+1)).get_data()

    temp_sel_opt=data[22]['选择的选项'].split(" ")


    sel_opt=[int(i) for i  in  temp_sel_opt]

    print("sel_opt=",sel_opt)
    print("len(sel_opt=)", len(sel_opt))

    options = ['-fgcse-after-reload', '-finline-functions', '-fipa-cp-clone', '-floop-interchange',
               '-floop-unroll-and-jam', '-fpeel-loops', '-fpredictive-commoning', '-fsplit-loops',
               '-fsplit-paths', '-ftree-loop-distribute-patterns', '-ftree-loop-distribution', '-ftree-loop-vectorize',
               '-ftree-partial-pre', '-ftree-slp-vectorize',
               '-funswitch-loops', '-fversion-loops-for-strides', '-fno-peephole2', '-ffast-math',
               '-fno-schedule-insns2', '-fno-caller-saves', '-funroll-all-loops',
               '-fno-inline-small-functions', '-finline-functions', '-fno-math-errno', '-fno-tree-pre', '-ftracer',
               '-fno-reorder-functions', '-fno-dce', '-fipa-cp-clone',
               '-fno-move-loop-invariants', '-fno-regmove', '-funsafe-math-optimizations', '-fno-tree-loop-optimize',
               '-fno-merge-constants', '-fno-omit-frame-pointer', '-fno-align-labels',
               '-fno-tree-ter', '-fno-tree-dse', '-fwrapv', '-fgcse-after-reload', '-fno-align-jumps',
               '-fno-asynchronous-unwind-tables', '-fno-cse-follow-jumps', '-fno-ivopts',
               '-fno-guess-branch-probability',
               '-fprefetch-loop-arrays', '-fno-tree-coalesce-vars', '-fno-common', '-fpredictive-commoning',
               '-fno-unit-at-a-time', '-fno-cprop-registers', '-fno-early-inlining', '-fno-delete-null-pointer-checks',
               '-fselective-scheduling2', '-fno-gcse', '-fno-inline-functions-called-once', '-funswitch-loops',
               '-fno-tree-vrp', '-fno-tree-dce', '-fno-jump-tables', '-ftree-vectorize', '-fno-argument-alias',
               '-fno-schedule-insns', '-fno-branch-count-reg', '-fno-tree-switch-conversion', '-fno-auto-inc-dec',
               '-fno-crossjumping', '-fno-tree-fre', '-fno-tree-reassoc', '-fno-align-functions', '-fno-defer-pop',
               '-fno-optimize-register-move', '-fno-strict-aliasing', '-fno-rerun-cse-after-loop', '-fno-tree-ccp',
               '-fno-ipa-cp', '-fno-if-conversion2', '-fno-tree-sra', '-fno-expensive-optimizations',
               '-fno-tree-copyrename', '-fno-ipa-reference', '-fno-ipa-pure-const', '-fno-thread-jumps',
               '-fno-if-conversion', '-fno-reorder-blocks', '-falign-loops']


    print("len(options)=",len(options))



    print_opt=[]

    for i in range(len(options)):
        if sel_opt[i+1]==1:
            print_opt.append(options[i])


    print(casename)
    print("第二阶段应使用的选项",print_opt)





