from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import random
import xlwt

# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
# fuzzy = [71, 0.4, 0.3, 0.3, 0.3, 0.8, 0.3, 0.7, 0.3, 0.4, 0.3, 0.1, 0.2, 0.4, 0.3, 0.5, 0.2, 0.0, 0.4, 0.3, 0.2, 0.3,
#          0.5, 0.3, 0.3, 0.3, 0.3, 0.3, 0.0, 0.4, 0.3, 0.5, 0.4, 0.7, 0.3, 0.2, 0.4, 0.2, 0.4, 0.1, 0.8, 0.3, 0.3, 0.4,
#          0.2, 0.2, 0.4, 0.1, 0.2, 0.4, 0.4, 0.4, 0.2, 0.2, 0.6, 0.4, 0.4, 0.4, 0.2, 0.2, 0.2, 0.2, 0.4, 0.3, 0.3, 0.6,
#          0.3, 0.4, 0.6, 0.3, 0.3, 0.3]


class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

#双向挖掘预选算法   输入为实验结果测试.xlsx 里面的数据为一个案例对应的50个采样解
if __name__ == '__main__':
    print("**************start*******************")

    casename="office_rsynth"
    for alg_temp_it in range(1):
        data = doExcel("一阶段采样集.xlsx", casename + str(alg_temp_it + 1)).get_data()
        print(data)
        #50个原始序列
        seqs = []
        #50个约简序列
        crc_seqs = []
        #50个目标值
        objs = []

        #保存excel数据到seqs、crc_seqs、objs
        for i in range(50):
            #保存50个原始序列到 seqs
            temp1 = data[i]['原始序列'].split()
            seq = [int(x) for x in temp1]
            seqs.append(seq)

            # 保存50个约简序列到 crc_seqs
            temp2 = data[i]['约简序列'].split()
            crc_seq = [int(x) for x in temp2]
            crc_seqs.append(crc_seq)

            # 保存50个目标值到 objs
            temp3 = data[i]['目标值']
            obj1 = float(temp3.strip('='))
            objs.append(obj1)

        print("seqs=",seqs)
        print("crc_seqs=",crc_seqs)
        print("objs=",objs)


        S= [86, 0.0, 0.0, 0.2, 0.0, 0.0, 0.4, 0.0, 0.0, 0.2, 0.0, 0.0, 0.2, 0.4, 0.2, 0.2, 0.0, 0.6, 1.0, 0.4, 0.0, 0.6, 0.0, 1.0, 0.0, 0.2, 0.8, 0.4, 0.2, 0.0, 0.4, 0.0, 0.0, 0.0, 0.6, 0.4, 0.0, 1.0, 0.8, 0.6, 0.8, 0.6, 0.2, 0.6, 0.4, 0.0, 0.8, 0.8, 0.8, 0.0, 1.0, 0.6, 0.2, 0.4, 0.0, 0.0, 0.0, 0.6, 1.0, 0.6, 0.8, 0.8, 0.0, 0.0, 0.0, 0.0, 0.0, 0.8, 0.2, 0.0, 0.8, 0.4, 0.0, 0.6, 0.4, 0.2, 0.0, 0.4, 0.0, 0.8, 0.0, 0.2, 0.6, 0.4, 1.0, 0.0, 0.0]
        NS=[]
        NS.append(86)

        #计算模糊假集
        for i in range(86):
            hhh=1-S[i+1]
            NS.append(hhh)

        sort_list = []
        sort_simpleList = []
        sort_time = []
        sorted_indices = np.argsort(objs, axis=0)
        # print(sorted_indices)
        # 从小到大排序
        for i in range(50):
            sort_list.append(seqs[sorted_indices[i]])
            sort_simpleList.append(crc_seqs[sorted_indices[i]])
            sort_time.append(objs[sorted_indices[i]])
        print("sort_List",sort_list)

        print("sort_simpleList=",sort_simpleList)
        print("sort_time=",sort_time)


        #用精简的序列集挖掘
        positive=[]
        positive.append(86)
        for i in range(86):
            positive.append(0)
        negative=[]
        negative.append(86)
        for i in range(86):
            negative.append(0)


        for i in range(25):
            positive_seq=sort_simpleList[i]
            negative_seq=sort_simpleList[i+25]

            for j in range(len(positive_seq)):
                if (positive_seq[j]==1):
                    positive[j+1]=positive[j+1]+1

            for j in range(len(negative_seq)):
                if (negative_seq[j]==1):
                    negative[j+1]=negative[j+1]+1


        print("positive=",positive)
        print("negative=",negative)

        excel_pos=[]
        excel_neg=[]
        excel_recallb=[]
        excel_recallw=[]
        excel_accuracy=[]
        selects_opt=[]
        for pos_num in range(10):
            for neg_num in range(10):
                # 使用精简的去选择选项
                temp_pos_list = []
                for hh in positive:
                    temp_pos_list.append(hh)
                temp_neg_list = []
                for hh in positive:
                    temp_neg_list.append(hh)

                pos=pos_num*2.5+2.5
                neg=neg_num*2.5+2.5
                excel_pos.append(pos/25)
                excel_neg.append(neg/25)
                select_opt=[]
                notselect_opt=[]

                select_opt.append(86)
                for i in range(86):
                    select_opt.append(0)

                notselect_opt.append(86)
                for i in range(86):
                    notselect_opt.append(0)
                for i in range(86):
                    if(temp_pos_list[i+1]>=pos ):
                        select_opt[i+1]=1
                    elif(temp_neg_list[i+1]<=neg and (temp_pos_list[i+1]!=0 or temp_neg_list[i+1]!=0)):
                        select_opt[i+1]=1
                selects_opt.append(select_opt)
                for i in range(86):
                    if (select_opt[i+1]==0):
                        notselect_opt[i+1]=1

                TP = 0.0
                FN = 0.0
                FP = 0.0
                TN = 0.0


                for i in range(86):
                    if(select_opt[i+1]==1):
                        TP = TP + S[i+1]
                        FP = FP + NS[i+1]
                for i in range(86):
                    if (notselect_opt[i+1]==1):
                        TN = TN + NS[i+1]
                        FN = FN + S[i+1]

                CRecallb = TP / (TP + FN)
                CRecallw = TN / (TN + FP)
                CAccuracy = (TP + TN) / (TP + TN + FP + FN)
                excel_recallb.append(CRecallb)
                excel_recallw.append(CRecallw)
                excel_accuracy.append(CAccuracy)

        print("excel_pos=",excel_pos)
        print("excel_neg=",excel_neg)
        print("selects_opt=",selects_opt)
        print("excel_recallb",excel_recallb)
        print("excel_recallw", excel_recallw)
        print("excel_accuracy", excel_accuracy)

        print("len,excel_pos=", len(excel_pos))
        print("len,excel_neg=", len(excel_neg))
        print("len,selects_opt=", len(selects_opt))
        print("len,excel_recallb", len(excel_recallb))
        print("len,excel_recallw", len(excel_recallw))
        print("len,excel_accuracy", len(excel_accuracy))

        workbook = xlwt.Workbook(encoding='utf-8')

        worksheet = workbook.add_sheet(casename)
        worksheet.write(0, 0, label='正向支持度')
        worksheet.write(0, 1, label='负向支持度')

        worksheet.write(0, 2, label='选择的选项')
        worksheet.write(0, 3, label='正向查全率')

        worksheet.write(0, 4, label='负向查全率')
        worksheet.write(0, 5, label='精度')

        for i in range(len(excel_pos)):
            worksheet.write(i+1, 0, label=str(excel_pos[i]))

        for i in range(len(excel_neg)):
            worksheet.write(i+1, 1, label=str(excel_neg[i]))

        for i in range(len(selects_opt)):
            list2=[str(i) for i in selects_opt[i]]
            list3=" ".join(list2)
            worksheet.write(i+1, 2, label=list3)

        for i in range(len(excel_pos)):
            worksheet.write(i+1, 3, label=str(excel_recallb[i]))

        for i in range(len(excel_pos)):
            worksheet.write(i+1, 4, label=str(excel_recallw[i]))

        for i in range(len(excel_pos)):
            worksheet.write(i+1, 5, label=str(excel_accuracy[i]))
        workbook.save(casename+str(alg_temp_it+1)+'.xlsx')
    print("**************end*******************")