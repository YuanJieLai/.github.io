from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
import random
import xlwt

# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False




class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

#双向挖掘预选算法   输入为实验结果测试.xlsx 里面的数据为一个案例对应的100个采样解
if __name__ == '__main__':

    print("**************start*******************")

    data = doExcel("D:\\paper\\毕业论文所需代码集合\\hlg\\绘图\\nussinov\\实验结果测试.xlsx", "nussinov7").get_data()
    obj=[]
    time1=[]
    for i in range(len(data)):
        obj.append(round(float(data[i]['目标值'][1:]),3))
        time1.append(round(float(data[i]['得到此目标值所需时间']),1))
    print(obj)

    mindata = obj[0]
    for i in range(len(obj)):
        if obj[i] < mindata:
            mindata = obj[i]
        else:
            obj[i] = mindata

    for i in range(len(obj)):
        print(time1[i],"    ",obj[i])






