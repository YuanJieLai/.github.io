from openpyxl import load_workbook  # 导入openpyxl
import numpy as np
import matplotlib.pyplot as plt
import pandas as pd
from sklearn.ensemble import AdaBoostRegressor
from sklearn.ensemble import GradientBoostingRegressor
from sklearn.ensemble import RandomForestRegressor
from sklearn.ensemble import RandomForestClassifier
from sklearn import svm
import xlwt

from sklearn.linear_model import LogisticRegression
from sklearn import tree
# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
# 处理Excel
plt.rcParams['font.sans-serif'] = ['SimHei']
plt.rcParams['axes.unicode_minus'] = False
class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data
class doExcel:
    def __init__(self, file_name, sheet_name):
        self.filename = file_name
        self.sheetname = sheet_name

    # 获取Excel中的方法
    def get_data(self):
        wb = load_workbook(self.filename)  # 打开Excel
        sheet = wb[self.sheetname]  # 定位表单
        test_data = []  # 创建一个空列表
        for row in range(2, sheet.max_row + 1):
            sub_data = {}
            for column in range(1, sheet.max_column + 1):
                sub_data[sheet.cell(1, column).value] = sheet.cell(row, column).value  # Excel的第一行数据作为字典的key；
            test_data.append(sub_data)  # 将每行的数据循环加到列表中
        return test_data

if __name__ == '__main__':

    casename="consumer_jpeg_c"

    resss=[]
    for e in range(1):

        data = doExcel("调优案例样本集.xlsx", casename + str(e + 1)).get_data()
        tmp_obj_list = []
        for i in range(len(data)):
            true_data1 = float(data[i]['目标值'])

            tmp_data = round(true_data1, 4)
            tmp_obj_list.append(tmp_data)
        #print(tmp_obj_list)




        tmp_ind_list = []
        for i in range(len(data)):
            true_data2 = data[i]['历史解']
            tmp_data = [int(x) for x in true_data2.split(" ")]
            tmp_ind_list.append(tmp_data)
        print("tmp_ind_list=",tmp_ind_list,"\nlen(tmp_ind_list)=",len(tmp_ind_list))

        tmp_train_x = []
        tmp_train_y = []


        for ite in range(10):
            print("************************",ite+1,"**************************")
            train_y = []


            for k in range(50):
                tmp_train_x.append(tmp_ind_list[ite*50*2+k])
           # print("len(tmp_train_x)=",len(tmp_train_x))

            for k in range(50):
                tmp_train_y.append(tmp_obj_list[ite*50*2+k])

            sort_train_y=sorted(tmp_train_y)
            # print("len(sort_train_y)",len(sort_train_y),"cur_sort_train_y=",sort_train_y)
            # print("len(tmp_train_x)",len(tmp_train_x),"tmp_train_x=",tmp_train_x)
            # print("len(tmp_train_y)",len(tmp_train_y),"tmp_train_y=", tmp_train_y)
            mid_y=sort_train_y[25*(ite+1)]
            mid_y=np.mean(sort_train_y)


            print("mid_y=",mid_y)
            pre_x=[]
            pre_y = []
            pre_true_y=[]
            for k in range(50):
                pre_x.append(tmp_ind_list[ite*50*2+50+k])
                pre_true_y.append(tmp_obj_list[ite*50*2+50+k])

            for y in range(len(tmp_train_y)):
                if tmp_train_y[y] < mid_y :
                    train_y.append(1)
                else :
                    train_y.append(0)
            #clf = RandomForestClassifier(n_estimators=20, random_state=0)
            #clf = svm.SVC()
            #clf = tree.DecisionTreeClassifier()
            clf = LogisticRegression()
            clf.fit(tmp_train_x, train_y)
            for data_x in pre_x:
                temp_model_x1 = np.array(data_x).reshape([1, np.array(data_x).shape[0]])
                pre_y.append(clf.predict(temp_model_x1)[0])
            true_p = 0.0
            print("pre_y=", pre_y)
            print("mid=", mid_y)
            print("pre_true_y", pre_true_y)
            for k in range(len(pre_true_y)):
                if pre_y[k] == 0:
                    if pre_true_y[k] >= mid_y:
                        true_p = true_p + 1
                else:
                    if pre_true_y[k] <= mid_y:
                        true_p = true_p + 1
            print("true_p=", true_p)
            resss.append(true_p)
    print("res=",resss)





